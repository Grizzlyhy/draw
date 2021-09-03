#coding=utf8

"""
界面随机挑选学号程序
可以选择随机数量，并打印在界面
"""


import sys
import random
from PyQt5.QtWidgets import QWidget, QPushButton, QApplication, QDesktopWidget, QLabel, QMessageBox,QFileDialog
from PyQt5.QtCore import QCoreApplication,QTimer
from PyQt5.Qt import QLineEdit, QFont
import openpyxl
import xlrd
import xlwt
import os
import time
import os.path
from random import sample

class Example(QWidget):

    def __init__(self):
        super().__init__()
        self.timer = None
        self.initUI()

    # 控制窗口显示在屏幕中心的方法
    def center(self):
        # 获得窗口
        qr = self.frameGeometry()
        # 获得屏幕中心点
        cp = QDesktopWidget().availableGeometry().center()
        # 显示到屏幕中心
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def  find_last_file(self):
        my_file = "D:/tmp/name.xls"
        if os.path.isfile(my_file) :
            self.filename="D:/tmp/name.xls"
            self.load_nameList()
        else :
            self._insert_name()


    def save_nameList(self):
        writebook = xlwt.Workbook()
        sheet = writebook.add_sheet('test')
        for i, name in enumerate(self.name_list):
            sheet.write(i,0,name)
        dirs = 'D:/tmp'
        if not os.path.exists(dirs):
            os.makedirs(dirs)
        try:
            writebook.save('D:/tmp/name.xls')
        except :
            print("储存失败")
        else :
            print("储存成功")
    def saveAns(self):
        writebook = xlwt.Workbook()
        sheet = writebook.add_sheet('test')
        for i, name in enumerate(self.temp_res):
            sheet.write(i,1,name)
            sheet.write(i,0,i+1)
        dirs = 'D:/tmp'
        if not os.path.exists(dirs):
            os.makedirs(dirs)
        try:
            writebook.save('D:/tmp/ans.xls')
        except :
            msg_box = QMessageBox(QMessageBox.Warning, '警告', '保存失败，请先关闭答案存储文件或截图保存结果')
            msg_box.exec_()
        else :
            self.save_ans.setEnabled(False)
            msg_box = QMessageBox(QMessageBox.Warning, '提示', '保存成功，请移步D:/tmp文件下查看结果')
            msg_box.exec_()

    def load_nameList(self):
        if self.filename.split('.')[1] == 'xlsx':
            inwb = openpyxl.load_workbook(self.filename)  # 读文件
            sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
            ws = inwb.get_sheet_by_name(sheetnames[0])
            rows = ws.max_row
            cols = ws.max_column
            flag=0
            for c in range(1,cols+1):
                if str(ws.cell(1,c).value).replace(' ', '') == "姓名" :
                    flag=c
            self.name_list=[]
            if flag == 0 :
                for r in range(1,rows+1):
                    self.name_list.append(str(ws.cell(r,1).value).strip())
            else :
                for r in range(2,rows+1):
                    self.name_list.append(str(ws.cell(r,flag).value).strip())
        else:
            readbook = xlrd.open_workbook(self.filename)
            sheet = readbook.sheet_by_index(0)
            nrows = sheet.nrows#行
            ncols = sheet.ncols#列
            flag=-1
            for c in range(0,ncols):
                if str(sheet.cell(0,c).value).replace(' ', '') == "姓名" :
                    flag=c
            self.name_list=[]
            if flag == -1 :
                for r in range(0,nrows):
                    self.name_list.append(str(sheet.cell(r,0).value).strip())
            else :
                for r in range(1,nrows):
                    self.name_list.append(str(sheet.cell(r,flag).value).strip())
            pass
        # self.name_list = list(filter(None,self.name_list))
        # print(self.name_list)
        self.index = len(self.name_list)-1
        print("插入成功")
        print(self.index)
        self.save_nameList()        
        pass


    def initUI(self):


        # create input textbox
        self.input = QLineEdit(self)
        self.input.move(300, 140)
        self.input.resize(50, 20)
        temp = QLabel(self)
        temp.move(230,140)
        temp.resize(60,20)
        temp.setText('抽签人数')

        # create presenting table
        self.textbox = QLabel(self)
        self.textbox.move(20, 20)
        self.textbox.resize(100, 300)
        self.textbox.setFont(QFont("Timers" , 28))

        self.textbox1 = QLabel(self)
        self.textbox1.move(20, 40)
        self.textbox1.resize(100, 300)
        self.textbox1.setFont(QFont("Timers" , 26))

        self.textbox2 = QLabel(self)
        self.textbox2.move(20, 60)
        self.textbox2.resize(100, 300)
        self.textbox2.setFont(QFont("Timers" , 26))

        self.textbox3 = QLabel(self)
        self.textbox3.move(20, 80)
        self.textbox3.resize(100, 300)
        self.textbox3.setFont(QFont("Timers" , 26))

        self.textbox4 = QLabel(self)
        self.textbox4.move(20, 100)
        self.textbox4.resize(100, 300)
        self.textbox4.setFont(QFont("Timers" , 26))

        self.textbox5 = QLabel(self)
        self.textbox5.move(20, 120)
        self.textbox5.resize(100, 300)
        self.textbox5.setFont(QFont("Timers" , 26))

        # # shuffle id list
        # self.name_list = list(map(str,range(42)))
        # random.shuffle(self.name_list)
   

        self.name_list=[]
        self.index = len(self.name_list)
        self.number = 1


        # find last file
        self.find_last_file()

        # create start, stop button
        self.start_ = QPushButton('开始', self)
        self.stop_ = QPushButton('结束', self)
        self.insert_name = QPushButton('插入名单',self)
        self.insert_name.move(215,210)
        self.save_ans = QPushButton('保存结果',self)
        self.save_ans.clicked.connect(self.saveAns)
        self.save_ans.move(315,210)
        self.start_.move(215, 170)
        self.stop_.move(315,170)
        self.start_.clicked.connect(self._start)
        self.stop_.clicked.connect(self._stop)
        self.insert_name.clicked.connect(self._insert_name)
        self.input.setText("1")
        self.save_ans.setEnabled(False)

        self.resize(600, 250)
        self.setWindowTitle('抽签程序')
        self.center()
        self.show()

    def _insert_name(self):
        openfile_name = QFileDialog.getOpenFileName(self,'选择学生名单文件','','Excel files(*.xlsx , *.xls)')
        if len(openfile_name[0].split('.')) <=1 :
            return
        if openfile_name[0].split('.')[1] == "xlsx" or openfile_name[0].split('.')[1] == "xls" :
            self.filename=openfile_name[0]
            self.load_nameList()
        else:
            pass


    def _range_name(self):
        if self.input.text():
            try :
                self.number = int(self.input.text())
            except :
                self.input.setText("1")
                self.number = int(self.input.text())
                msg_box = QMessageBox(QMessageBox.Warning, '警告', '请输入数字')
                msg_box.exec_()
                return
        temp_res1 =[]
        temp_res2 =[]
        temp_res3 =[]
        temp_res4 =[]
        temp_res5 =[]
        if self.number <= 0:
            self.timer.stop()
            msg_box = QMessageBox(QMessageBox.Warning, '警告', '请输入正确人数')
            msg_box.exec_()
            self.input.setText("1")
            self.number = int(self.input.text())
            return
        if self.number > len(self.name_list):
            self.timer.stop()
            msg_box = QMessageBox(QMessageBox.Warning, '警告', '选择的人数不能大于名单总人数')
            msg_box.exec_()
            return
        else :
            self.temp_res =sample(self.name_list, self.number)

        # if len(temp_res) < 4:
        #     self.textbox.setFont(QFont("Timers" , 28))
        # elif len(temp_res) >= 4:
        #     self.textbox.setFont(QFont("Timers" , 26))
        #     self.textbox1.setFont(QFont("Timers" , 26))
        # if len(temp_res) > 4:
        #     temp_res1 = temp_res[4:]
        #     temp_res = temp_res[:4]

        temp_res6 =self.temp_res[55:]
        temp_res5 = self.temp_res[44:55]
        temp_res4 = self.temp_res[33:44]
        temp_res3 = self.temp_res[22:33]
        temp_res2 = self.temp_res[11:22]
        temp_res1 = self.temp_res[:11]
        self.textbox.setFont(QFont("Timers" , 8))
        self.textbox1.setFont(QFont("Timers" , 8))
        self.textbox2.setFont(QFont("Timers" , 8))
        self.textbox3.setFont(QFont("Timers" , 8))
        self.textbox4.setFont(QFont("Timers" , 8))
        self.textbox5.setFont(QFont("Timers" , 8))
        self.textbox.setText(','.join(temp_res1))
        self.textbox.adjustSize()
        self.textbox1.setText(','.join(temp_res2))
        self.textbox1.adjustSize()
        self.textbox2.setText(','.join(temp_res3))
        self.textbox2.adjustSize()
        self.textbox3.setText(','.join(temp_res4))
        self.textbox3.adjustSize()
        self.textbox4.setText(','.join(temp_res5))
        self.textbox4.adjustSize()
        self.textbox5.setText(','.join(temp_res6))
        self.textbox5.adjustSize()
       

    def _start(self):
        if len(self.name_list) == 0:
            msg_box = QMessageBox(QMessageBox.Warning, '警告', '请先插入学生名单')
            msg_box.exec_()
            return
        if not self.input.text():
            msg_box = QMessageBox(QMessageBox.Warning, '警告', '请输入随机人数')
            msg_box.exec_()
            return
        

        if not self.timer:
            self.timer = QTimer()
        self.timer.timeout.connect(self._range_name)
        self.timer.start(50)

        if self.start_.isEnabled():
            self.start_.setEnabled(False)

        if self.insert_name.isEnabled():
            self.insert_name.setEnabled(False)
        #self.stop_.setEnabled(True)
        QApplication.processEvents()

    def _stop(self):
        self.save_ans.setEnabled(True)
        if not self.start_.isEnabled():
            self.start_.setEnabled(True)
        if not self.insert_name.isEnabled():
            self.insert_name.setEnabled(True)
        if self.timer:
            self.timer.stop()



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())
